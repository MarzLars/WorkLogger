import time
import os
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

class TimeTracker:
    def __init__(self):
        """
        Initialize the TimeTracker with default values.
        """
        self.start_time = None
        self.elapsed_time = 0
        self.running = False

    def start(self):
        """
        Start the time tracker.
        """
        if not self.running:
            self.start_time = time.time()
            self.running = True

    def _update_elapsed_time(self):
        """
        Update the elapsed time if the tracker is running.
        """
        if self.running:
            self.elapsed_time += time.time() - self.start_time
            self.start_time = time.time()

    def pause(self):
        """
        Pause the time tracker.
        """
        if self.running:
            self.elapsed_time += time.time() - self.start_time
            self.running = False

    def stop(self):
        """
        Stop the time tracker and return the elapsed time.
        """
        if self.running:
            self.elapsed_time += time.time() - self.start_time
            self.running = False

    def reset(self):
        """
        Reset the time tracker to its initial state.
        """
        self.start_time = None
        self.elapsed_time = 0
        self.running = False

    def get_elapsed_time(self):
        if self.running:
            return self.elapsed_time + (time.time() - self.start_time)
        return self.elapsed_time

    def _apply_header_style(self, sheet):
        """
        Apply styles to the header row of the Excel sheet.
        """
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        dark_grey_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
        bold_font = Font(name='Arial', size=8, bold=True)
        
        for cell in sheet[1]:
            if cell.value in ["Hours", "Minutes", "Seconds"]:
                cell.fill = orange_fill
            else:
                cell.fill = dark_grey_fill
            cell.font = bold_font

    def _apply_row_style(self, sheet, row_number):
        """
        Apply styles to a specific row of the Excel sheet.
        """
        regular_font = Font(name='Arial', size=8)
        
        for cell in sheet[row_number]:
            cell.font = regular_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        sheet[f'A{row_number}'].alignment = Alignment(wrap_text=True)
        date_cell = sheet[f'B{row_number}']
        date_cell.number_format = 'DD/MM/YYYY'

    def _adjust_column_widths(self, sheet):
        """
        Adjust the column widths of the Excel sheet.
        """
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            max_length = 0
            column = sheet[col]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[col].width = adjusted_width
        
        if sheet.column_dimensions['B'].width < 12:
            sheet.column_dimensions['B'].width = 12

    def log_time(self, description, file_path='time_log.xlsx', csv_file_path='time_log.csv'):
        """
        Log the elapsed time to both Excel and CSV files.
        """
        # Ensure log files exist
        self.ensure_log_file_exists(file_path, csv_file_path)

        # Calculate hours, minutes, and seconds
        hours, remainder = divmod(self.elapsed_time, 3600)
        minutes, seconds = divmod(remainder, 60)

        # Log to CSV
        current_date = datetime.now().strftime("%Y-%m-%d")
        week_number = datetime.now().isocalendar()[1]
        self.log_time_to_csv(description, current_date, week_number, hours, minutes, seconds, csv_file_path)

        # Log to Excel
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
        else:
            workbook = Workbook()
        sheet = workbook.active
        sheet.append([description, current_date, week_number, self.elapsed_time, hours, minutes, seconds])
        workbook.save(file_path)

    def log_time_to_csv(self, description, current_date, week_number, hours, minutes, seconds, csv_file_path):
        """
        Log the elapsed time to a CSV file.
        """
        file_exists = os.path.isfile(csv_file_path)
        with open(csv_file_path, mode='a', newline='') as file:
            writer = csv.writer(file, delimiter=';')
            if not file_exists:
                file.write("sep=;\n")
                writer.writerow(['Description', 'Date', 'Week', 'Time Spent', 'Hours', 'Minutes', 'Seconds'])
            writer.writerow([description, current_date, week_number, hours + (minutes / 60), int(hours), int(minutes), int(seconds)])

    def ensure_log_file_exists(self, file_path='time_log.xlsx', csv_file_path='time_log.csv'):
        """
        Ensure that the log files exist, creating them if necessary.
        """
        if not os.path.exists(file_path):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(['Description', 'Date', 'Week', 'Time Spent', 'Hours', 'Minutes', 'Seconds'])
            self._apply_header_style(sheet)
            self._adjust_column_widths(sheet)
            workbook.save(file_path)
        
        if not os.path.exists(csv_file_path):
            with open(csv_file_path, mode='w', newline='') as file:
                writer = csv.writer(file, delimiter=';')
                file.write("sep=;\n")
                writer.writerow(['Description', 'Date', 'Week', 'Time Spent', 'Hours', 'Minutes', 'Seconds'])